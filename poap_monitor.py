"""

POAP Migration Monitor

======================

Reads Outlook email chains for each SAP migration object, detects the current

activity via Dify workflow (GAIN AI) or keyword fallback, compares it against

the POAP planned timeline, and writes a RAG-status dashboard (Excel + HTML).

Usage

-----

  python poap_monitor.py           # continuous polling (every 5 min)

  python poap_monitor.py --once    # single run then exit

  python poap_monitor.py --interval 600   # every 10 minutes

"""

import os, re, json, time, html as html_lib, argparse

from concurrent.futures import ThreadPoolExecutor, as_completed

from datetime import datetime, date, timedelta

import openpyxl

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from openpyxl.utils import get_column_letter

# win32com imported lazily inside the email reader so the file can be

# imported and tested on machines without Outlook / pywin32.


# ═══════════════════════════════════════════════════════════════════════════════

#  SECTION 1 — CONFIG

# ═══════════════════════════════════════════════════════════════════════════════

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

POAP_FILE            = os.path.join(BASE_DIR, "LDC_R4.0_Mock2_EU_LoadPlan.xlsx")

POAP_SHEET           = "2. Load Plan-EU"          # name of the POAP tab in the Excel file

EMAIL_SUBJECT_PREFIX = "LDC R4.0 : MOCK2 : EU"

MAX_EMAILS_PER_CHAIN = 10

SCAN_ROOT_FOLDERS    = ["Inbox", "Archive","R4 Mock2 EU"]

LOOKBACK_DAYS        = 90

CACHE_OVERLAP_MINS   = 10       # safety overlap on incremental scans

POLL_SECONDS         = 60       # continuous mode interval

DUE_HOURS            = 3        # hours after last email before status becomes Due

OUTPUT_EXCEL     = os.path.join(BASE_DIR, "POAP_Status_Dashboard.xlsx")

OUTPUT_HTML      = os.path.join(BASE_DIR, "dashboard.html")

DASHBOARD_CONFIG = os.path.join(BASE_DIR, "dashboard.config.json")

STATE_DIR        = os.path.join(BASE_DIR, "state")

STATE_FILE       = os.path.join(STATE_DIR, "monitor_state.json")

EMAIL_CACHE_FILE = os.path.join(STATE_DIR, "email_cache.json")

# ── AI / LLM config ───────────────────────────────────────────────────────────

# To disable Dify calls and use keyword scoring only → set LLM_ENABLED = False

LLM_ENABLED = True

LLM_TIMEOUT = 30

from dotenv import load_dotenv

load_dotenv()

GAIN_API_URL_ANALYSER = os.getenv("GAIN_API_URL_ANALYSER")

GAIN_API_KEY_ANALYSER = os.getenv("GAIN_API_KEY_ANALYSER")

GAIN_USER_ID          = os.getenv("GAIN_USER_ID", "poap-monitor")


# ═══════════════════════════════════════════════════════════════════════════════

#  SECTION 2 — POAP PARSER

# ═══════════════════════════════════════════════════════════════════════════════

ACTIVITY_ORDER = ['Extract', 'Pre-T', 'SIF', 'Transform', 'Preload-Val', 'Load', 'Post', 'REPL']

TASK_GROUP_MAP = {

    'Extract': 'Extract', 'Pre-T': 'Pre-T', 'SIF': 'SIF',

    'Transform': 'Transform', 'Pre': 'Preload-Val',

    'Load': 'Load', 'Load1': 'Load', 'Load2': 'Load', 'MANUAL': 'Load',

    'Post': 'Post', 'REPL': 'REPL',

}


def _to_date(v):

    if isinstance(v, datetime): return v.date()

    if isinstance(v, date):     return v

    return None


_poap_cache = None

_poap_mtime = None

def load_objects():

    """Return dict {obj_id: {...}}. Re-reads Excel only when the file changes."""

    global _poap_cache, _poap_mtime

    try:

        mtime = os.path.getmtime(POAP_FILE)

    except OSError:

        mtime = None

    if _poap_cache is not None and mtime == _poap_mtime:

        print('      (POAP plan unchanged — using cached data)')

        return _poap_cache

    wb      = openpyxl.load_workbook(POAP_FILE, read_only=True, data_only=True)

    objects = {}

    if POAP_SHEET not in wb.sheetnames:

        raise KeyError(

            f"Sheet '{POAP_SHEET}' not found in {POAP_FILE}.\n"

            f"Available sheets: {wb.sheetnames}\n"

            f"Update POAP_SHEET at the top of poap_monitor.py to match the correct tab name."

        )

    for i, row in enumerate(wb[POAP_SHEET].iter_rows(values_only=True)):

        if i < 6: continue

        oid = row[1]

        if not oid or not isinstance(oid, str) or oid in ('ID', 'FIN'): continue

        objects[oid] = {

            'name':            str(row[2]).strip() if row[2] else oid,

            'stream':          str(row[6]).strip() if row[6] else '',

            'owner_gdme':      str(row[3]).strip() if row[3] else '',

            'owner_eu':        str(row[4]).strip() if row[4] else '',

            'start_week':      str(row[9]).strip() if row[9] else '',

            'activities':      {},

            'overall_status':  'Unknown',

            'pct_complete':    0.0,

            'load_cycle_step': '',

        }

    load_plan_sheet    = next((s for s in wb.sheetnames if '2. Load Plan' in s), None)

    load_summary_sheet = next((s for s in wb.sheetnames if '3. Load Object Summary' in s), None)

    if not load_plan_sheet:

        print('  WARNING: No sheet matching "2. Load Plan" found in Excel.')

    if not load_summary_sheet:

        print('  WARNING: No sheet matching "3. Load Object Summary" found in Excel.')

    col = {}

    first = True

    for row in wb[load_plan_sheet].iter_rows(values_only=True) if load_plan_sheet else []:

        if first:

            first = False

            headers = [str(c).strip().upper() if c else '' for c in row]

            col = {h: i for i, h in enumerate(headers) if h}

            missing = [n for n in ('COUNTRY', 'OBJECT ID', 'TASKGROUP', 'DDME',
                                   'START BASELINE', 'FINISH BASELINE', 'STATUS') if n not in col]

            if missing:

                print(f'  WARNING: Load Plan sheet missing expected columns: {missing}')

            continue

        country = row[col['COUNTRY']]       if 'COUNTRY'        in col else row[1]

        oid     = row[col['OBJECT ID']]     if 'OBJECT ID'      in col else row[2]

        tg      = row[col['TASKGROUP']]     if 'TASKGROUP'      in col else None

        if not oid or oid not in objects: continue

        if 'DDME' in col and row[col['DDME']]:

            objects[oid]['owner_eu'] = str(row[col['DDME']]).strip()

        if country != 'ALL' or not tg: continue

        canon = TASK_GROUP_MAP.get(str(tg))

        if not canon: continue

        p_start  = _to_date(row[col['START BASELINE']])  if 'START BASELINE'  in col else None

        p_finish = _to_date(row[col['FINISH BASELINE']]) if 'FINISH BASELINE' in col else None

        status_txt = str(row[col['STATUS']]).strip() if 'STATUS' in col and row[col['STATUS']] else ''

        existing = objects[oid]['activities'].get(canon)

        if existing:

            if p_finish and (not existing['planned_finish'] or p_finish > existing['planned_finish']):

                existing['planned_finish'] = p_finish

                existing['poap_status']    = status_txt

        else:

            objects[oid]['activities'][canon] = {

                'planned_start': p_start, 'planned_finish': p_finish, 'poap_status': status_txt,

            }

    EXEC_STATUSES = {'Complete', 'At Risk', 'On Track', 'Not Started', 'Not Relevant'}

    LOAD_STEPS    = {'Completed', 'Loading', 'SIF Reviewing', 'Pre-verifying',

                     'Transforming', 'Pre-Transforming', 'Constructing',

                     'Replication', 'Not Started', 'Post-verifying'}

    for i, row in enumerate(wb[load_summary_sheet].iter_rows(values_only=True) if load_summary_sheet else []):

        if i < 3: continue

        oid = exec_stat = load_step = None; pct = None

        for cell in row:

            if cell is None: continue

            cs = str(cell).strip()

            if cs in objects:                              oid       = cs

            elif isinstance(cell,(int,float)) and 0<=float(cell)<=1: pct = float(cell)

            elif cs in EXEC_STATUSES:                      exec_stat = cs

            elif cs in LOAD_STEPS:                         load_step = cs

        if oid:

            if exec_stat: objects[oid]['overall_status']  = exec_stat

            if pct is not None: objects[oid]['pct_complete'] = round(float(pct), 4)

            if load_step: objects[oid]['load_cycle_step'] = load_step

    wb.close()

    _poap_cache = objects

    _poap_mtime = mtime

    return objects


def get_expected_activity(obj_data, today=None):

    if today is None: today = date.today()

    acts = obj_data.get('activities', {})

    has_date = False

    for act in ACTIVITY_ORDER:

        if act not in acts: continue

        finish = acts[act].get('planned_finish')

        if not finish: continue

        has_date = True

        if finish >= today: return act

    return 'Complete' if has_date else 'Unknown'
 
 
def get_planned_finish(obj_data, activity):

    return obj_data.get('activities', {}).get(activity, {}).get('planned_finish')
 
 
# ═══════════════════════════════════════════════════════════════════════════════

#  SECTION 3 — EMAIL CACHE

# ═══════════════════════════════════════════════════════════════════════════════
 
def _email_to_dict(e):

    d = dict(e)

    if isinstance(d.get('date'), datetime): d['date'] = d['date'].isoformat()

    return d
 
def _dict_to_email(d):

    e = dict(d)

    if isinstance(e.get('date'), str): e['date'] = datetime.fromisoformat(e['date'])

    return e
 
 
def cache_load():

    """Returns (last_scan_dt, chains, detections). last_scan_dt is None on first run."""

    if not os.path.exists(EMAIL_CACHE_FILE): return None, {}, {}

    try:

        with open(EMAIL_CACHE_FILE, encoding='utf-8') as f: raw = json.load(f)

        last_scan = datetime.fromisoformat(raw['last_scan_utc']) if raw.get('last_scan_utc') else None

        chains    = {oid: [_dict_to_email(e) for e in emails]

                     for oid, emails in raw.get('chains', {}).items()}

        detections = {}

        for oid, det in raw.get('detections', {}).items():

            d = dict(det)

            if d.get('last_email_date'):

                try: d['last_email_date'] = datetime.fromisoformat(d['last_email_date'])

                except: d['last_email_date'] = None

            detections[oid] = d

        return last_scan, chains, detections

    except Exception as exc:

        print(f'  Warning: could not read email cache ({exc}). Falling back to full scan.')

        return None, {}, {}
 
 
def cache_save(chains, detections=None):

    os.makedirs(STATE_DIR, exist_ok=True)

    def _serial_det(det):

        d = dict(det)

        if isinstance(d.get('last_email_date'), datetime):

            d['last_email_date'] = d['last_email_date'].isoformat()

        return d

    payload = {

        'last_scan_utc': datetime.now().isoformat(),

        'chains': {oid: [_email_to_dict(e) for e in emails] for oid, emails in chains.items()},

        'detections': {oid: _serial_det(det) for oid, det in (detections or {}).items()},

    }

    with open(EMAIL_CACHE_FILE, 'w', encoding='utf-8') as f:

        json.dump(payload, f, indent=2, ensure_ascii=False)
 
 
def cache_merge(cached_chains, new_emails):

    """Merge new emails into cached chains; deduplicate; cap at MAX_EMAILS_PER_CHAIN."""

    merged = dict(cached_chains)

    for oid, new_list in new_emails.items():

        existing = merged.get(oid, [])

        def _key(e):

            dt = e['date']

            return (

                e.get('subject','').strip().lower(),

                e.get('sender', '').strip().lower(),

                dt.replace(second=0, microsecond=0) if isinstance(dt, datetime) else dt,

            )

        seen     = {_key(e) for e in existing}

        combined = [e for e in new_list if _key(e) not in seen] + existing

        combined.sort(key=lambda e: e['date'], reverse=True)

        merged[oid] = combined[:MAX_EMAILS_PER_CHAIN]

    return merged
 
 
def cache_get_cutoff(last_scan_dt):

    """Incremental cutoff (last run − overlap) or full-scan cutoff on first run."""

    if last_scan_dt is None:

        return datetime.now() - timedelta(days=LOOKBACK_DAYS)

    return last_scan_dt - timedelta(minutes=CACHE_OVERLAP_MINS)
 
 
# ═══════════════════════════════════════════════════════════════════════════════

#  SECTION 4 — EMAIL READER  (requires Outlook + pywin32)

# ═══════════════════════════════════════════════════════════════════════════════
 
_FOLDER_IDS = {'Inbox': 6, 'Archive': 23, 'Sent Items': 5, 'Deleted Items': 3}
 
def _strip_html(raw):

    if not raw: return ''

    text = re.sub(r'<(script|style)[^>]*>.*?</(script|style)>', '', raw,

                  flags=re.DOTALL | re.IGNORECASE)

    text = re.sub(r'<[^>]+>', ' ', text)

    return re.sub(r'\s+', ' ', html_lib.unescape(text)).strip()
 
def _get_body(item):

    try:

        b = item.Body

        if b and b.strip(): return b.strip()

    except Exception: pass

    try: return _strip_html(item.HTMLBody)

    except Exception: return ''
 
def _com_dt(rt):

    return datetime(rt.year, rt.month, rt.day, rt.hour, rt.minute, rt.second)
 
def _extract_obj_id(subject, known_ids):

    subj_upper = subject.upper()

    for kid in sorted(known_ids, key=len, reverse=True):

        if (f': {kid} ' in subj_upper or f': {kid}\u2013' in subj_upper

                or f': {kid}-' in subj_upper):

            return kid

    m = re.search(r'EU\s*:\s*([A-Z][A-Z0-9]{0,5})', subject, re.IGNORECASE)

    if m:

        c = m.group(1).upper()

        return c if 1 <= len(c) <= 6 else None

    return None
 
def _body_obj_id(body, known_ids):

    body_upper = body[:600].upper()

    best_oid, best_pos = None, len(body_upper) + 1

    for kid in known_ids:

        pos = body_upper.find(kid.upper())

        if pos != -1 and pos < best_pos:

            best_pos, best_oid = pos, kid

    return best_oid
 
def _scan_folder(folder, cutoff, prefix_upper, known_ids, bucket):

    try:

        all_items = folder.Items

        date_str  = cutoff.strftime('%m/%d/%Y %I:%M %p')

        try:

            items = all_items.Restrict(f"[ReceivedTime] >= '{date_str}'")

        except Exception:

            items = all_items

        for item in items:

            try:

                if item.Class != 43: continue

                rt_naive = _com_dt(item.ReceivedTime)

                if rt_naive < cutoff: continue

                subj = str(item.Subject or '')

                if prefix_upper not in subj.upper(): continue

                subj_oid = _extract_obj_id(subj, known_ids)

                if not subj_oid:

                    body_oid = _body_obj_id(_get_body(item), known_ids)

                    if not body_oid: continue

                    subj_oid = body_oid

                body = _get_body(item)

                oid = subj_oid

                bucket.setdefault(oid, []).append({

                    'subject': subj,

                    'sender':  str(getattr(item,'SenderName','') or

                                   getattr(item,'SenderEmailAddress','') or ''),

                    'date':    rt_naive,

                    'body':    body,

                })

            except Exception:

                continue

    except Exception:

        pass

    try:

        for sf in folder.Folders: _scan_folder(sf, cutoff, prefix_upper, known_ids, bucket)

    except Exception:

        pass
 
def get_new_emails(known_ids=None, since_dt=None):

    """Scan ALL Outlook folders/subfolders from since_dt onward. Returns {obj_id: [emails]} newest-first."""

    import win32com.client

    if known_ids is None: known_ids = set()

    cutoff       = since_dt or (datetime.now() - timedelta(days=LOOKBACK_DAYS))

    prefix_upper = EMAIL_SUBJECT_PREFIX.upper()

    namespace    = win32com.client.Dispatch('Outlook.Application').GetNamespace('MAPI')

    bucket       = {}

    try:

        for store in namespace.Stores:

            try:

                for folder in store.GetRootFolder().Folders:

                    _scan_folder(folder, cutoff, prefix_upper, known_ids, bucket)

            except Exception as e:

                print(f'  Warning: could not scan store "{getattr(store, "DisplayName", "?")}": {e}')

    except Exception:

        # Fallback: use namespace.Folders if Stores is unavailable

        for fname in SCAN_ROOT_FOLDERS:

            fid = _FOLDER_IDS.get(fname)

            try:

                root = namespace.GetDefaultFolder(fid) if fid else None

                if root: _scan_folder(root, cutoff, prefix_upper, known_ids, bucket)

            except Exception as e:

                print(f'  Warning: could not scan "{fname}": {e}')

    for oid in bucket:

        bucket[oid].sort(key=lambda e: e['date'], reverse=True)

    return bucket

 
 
# ═══════════════════════════════════════════════════════════════════════════════

#  SECTION 5 — ACTIVITY DETECTOR

# ═══════════════════════════════════════════════════════════════════════════════
 
ACTIVITY_KEYWORDS = {

    'Preload-Val': [

        'preload validation', 'pre-load validation', 'preload check',

        'pre-load check', 'preload validations', 'preload val',

        'pre load validation', 'preload activities', 'preload review',

        'no re-transformation', 'no retransformation', 'pre-load activities',

    ],

    'Transform': [

        'transform', 'transformation', 'transformed', 'transforming',

        'etl run', 'transform complete', 'transform done',

        'transformation complete', 'data transformation',

        'transform output', 'transform file', 'transform started',

        'running transform', 'transformation started',

    ],

    'Load': [

        'load', 'loaded', 'loading', 'gmdf',

        'load complete', 'data loaded', 'load done',

        'load successful', 'load started', 'load in progress',

        'load 1', 'load 2', 'loading data', 'data load',

        'ldc load', 'load execution', 'load initiated',

        'upload', 'run again the upload', 're-upload',

    ],

    'Post': [

        'post load', 'post-load', 'post check', 'post verification',

        'reconciliation', 'post activities', 'post load check',

        'post load validation', 'post load review',

        'post-load verification', 'post activities complete',

        'post-load activities',

    ],

    'REPL': [

        'replication', 'repl', 'replicated', 'replicating',

        'delta load', 'replication done', 'replication complete',

        'replicate', 'replication started', 'delta replication',

    ],

    'Complete': [

        'fully complete', 'object complete', 'migration complete',

        'all activities complete', 'wrap up', 'wrapped up',

        'all steps complete', 'object closed', 'migration closed',

    ],

}
 
COMPLETION_WORDS = [

    'complete', 'completed', 'done', 'delivered', 'approved',

    'signed off', 'closed', 'finished', 'successful', 'ok',

    'confirmed', 'received', 'ready', 'passed', 'accepted',

]
 
REQUEST_KEYWORDS = [

    'requested', 'requesting', 'please proceed', 'please perform',

    'please run', 'please load', 'please transform', 'kindly proceed',

    'kindly perform', 'proceed with', 'initiate load', 'initiate transform',

    'step2 requested', 'dvt step2', 'dvt step 2', 'correction load',

    'reload requested', 'correction file', 'request to perform',

    'request to run', 'request to load', 'could you please', 'can you please', 'please',

]
 
REQUEST_OVERRIDE_KEYWORDS = [kw for kw in REQUEST_KEYWORDS if kw.strip().lower() != 'please']
 
BLOCKER_KEYWORDS = [

    'blocked', 'on hold', 'waiting for', 'pending approval',

    'not received', 'data missing', 'delayed by', 'causing delay',

    'has an issue', 'issue with', 'issue found',

    'error found', 'load failed', 'load error', 'transform failed',

    'failure in', 'rejected by', 'rejection from', 'escalat',

    'cannot proceed', 'cannot load', 'unable to',

    'approval needed', 'awaiting approval', 'no response',

    'not responded', 'unresolved', 'still waiting',

]
 
_DETECT_ORDER = ['Transform', 'Preload-Val', 'Load', 'Post', 'REPL']
 
_REPLY_SEP = re.compile(

    r'(-----+\s*Original Message\s*-----+'

    r'|From:.*?Sent:.*?To:.*?Subject:'

    r'|On .{5,80} wrote:'

    r'|_{5,}'

    r'|\d{1,2}/\d{1,2}/\d{2,4}\s+\d{1,2}:\d{2})',

    re.IGNORECASE | re.DOTALL,

)
 
_HEADER_LINE = re.compile(

    r'^(from|to|sent|cc|bcc|subject|date|contact|tel|phone|mobile)\s*:', re.IGNORECASE

)

_SIGNATURE_START = re.compile(

    r'^(regards|thanks|thank you|best regards|cheers|sincerely|hi all|hi,)\b',

    re.IGNORECASE

)
 
def _strip_reply(body):

    m = _REPLY_SEP.search(body)

    return body[:m.start()].strip() if m else body.strip()
 
def _kw_score(text, keywords):

    return sum(1 for kw in keywords if kw in text)
 
_SENTENCE_SPLIT = re.compile(r'(?<=[.!?])\s+')
 
def _summarise_text(text: str, max_sentences: int = 3, max_chars: int = 360) -> str:

    if not text:

        return ''

    sentences = _SENTENCE_SPLIT.split(text.strip())

    picked = []

    for sent in sentences:

        sent = sent.strip()

        if not sent:

            continue

        picked.append(sent)

        if len(picked) >= max_sentences:

            break

    summary = ' '.join(picked)

    if len(summary) > max_chars:

        summary = summary[:max_chars].rsplit(' ', 1)[0] + '…'

    return summary
 
def _guess_activity_from_text(text: str | None) -> str | None:

    if not text:

        return None

    lower = text.lower()

    best_act, best_score = None, 0

    for act in ('Transform', 'Preload-Val', 'Load', 'Post', 'REPL'):

        kws = ACTIVITY_KEYWORDS.get(act, [])

        score = sum(1 for kw in kws if kw in lower)

        if score > best_score:

            best_act, best_score = act, score

    return best_act if best_score else None
 
def _analyse_latest_text(text: str):

    if not text:

        return None, False, False

    lower = text.lower()

    activity_hint = _guess_activity_from_text(text)

    is_request    = any(kw in lower for kw in REQUEST_OVERRIDE_KEYWORDS)

    is_completed  = any(cw in lower for cw in COMPLETION_WORDS)

    return activity_hint, is_request, is_completed
 
def _get_email_body_for_llm(email, max_chars=800):

    raw = email.get('body', '')

    new = _strip_reply(raw).strip()

    if len(new) > 80:

        return new

    source = new if len(new) > 10 else raw

    paragraphs = re.split(r'\n\s*\n', source)

    meaningful = []

    total = 0

    for para in paragraphs:

        lines = [l.strip() for l in para.splitlines() if l.strip()]

        if lines and all(_HEADER_LINE.match(l) for l in lines):

            continue

        text = re.sub(r'\s+', ' ', para).strip()

        if len(text) < 20:

            continue

        if _SIGNATURE_START.match(text) and len(text) < 160:

            continue

        meaningful.append(text)

        total += len(text)

        if total >= max_chars:

            break

    return ' | '.join(meaningful) if meaningful else source
 
def _detect_via_llm(emails):

    import requests, urllib3

    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
 
    latest_email = emails[0]

    body         = _get_email_body_for_llm(latest_email)

    synopsis     = _summarise_text(body) or body[:360]
 
    chain_text = (

        "Newest email only:\n"

        f"Date   : {latest_email.get('date', '')}\n"

        f"Sender : {latest_email.get('sender', '')}\n"

        f"Summary: {synopsis}"

    )
 
    headers = {

        "Authorization": f"Bearer {GAIN_API_KEY_ANALYSER}",

        "Content-Type":  "application/json",

    }

    payload = {

        "inputs":        {"query": chain_text},

        "response_mode": "blocking",

        "user":          GAIN_USER_ID,

    }
 
    resp   = requests.post(GAIN_API_URL_ANALYSER, json=payload, headers=headers,

                           timeout=LLM_TIMEOUT, verify=False)

    resp.raise_for_status()

    result = json.loads(resp.json()["data"]["outputs"]["Output"])
 
    cur_act    = result.get('current_activity', 'Unknown')

    raw_status = result.get('activity_status', 'Unknown')

    llm_summary = result.get('summary', '').strip() or synopsis[:200]
 
    if isinstance(raw_status, str) and raw_status.endswith(' Requested'):

        raw_status = 'Requested'
 
    if raw_status == 'Requested' and cur_act not in ('Unknown', None):

        act_status = f"{cur_act} Requested"

    else:

        act_status = raw_status
 
    act_hint, is_request, is_completed = _analyse_latest_text(synopsis)
 
    if is_request and not is_completed:

        cur_act = cur_act if cur_act not in ('Unknown', None) else (act_hint or 'Load')

        act_status = f"{cur_act} Requested"

    elif is_completed:

        if cur_act in ('Unknown', None) and act_hint:

            cur_act = act_hint

        if act_status not in ('Completed', 'Completed with Errors'):

            act_status = 'Completed'

    elif cur_act in ('Unknown', None) and act_hint:

        cur_act = act_hint
 
    dt = latest_email.get('date')

    stamp = dt.strftime('%d-%b-%Y') if isinstance(dt, datetime) else str(dt or '')

    combined_summary = f"[{stamp}] {synopsis}"
 
    return {

        'current_activity': cur_act,

        'activity_status':  act_status,

        'summary':          combined_summary,

        'blockers':         result.get('blockers'),

        'last_email_date':   latest_email['date'],

        'last_email_sender': latest_email.get('sender', ''),

        'email_count':       len(emails),

    }

 

def _detect_via_keywords(emails):

    if not emails:

        return {'current_activity': 'Unknown', 'activity_status': 'Unknown',

                'summary': 'No emails found for this object.',

                'blockers': None, 'last_email_date': None,

                'last_email_sender': '', 'email_count': 0}
 
    email = emails[0]

    cleaned = _strip_reply(email.get('subject','') + ' ' + email.get('body','')).lower()
 
    scores = {act: _kw_score(cleaned, kws) for act, kws in ACTIVITY_KEYWORDS.items()}

    top_act = max(scores, key=scores.get)

    if scores[top_act] == 0:

        top_act = 'Unknown'
 
    activity_completed = top_act in _DETECT_ORDER and any(

        cw in cleaned for cw in COMPLETION_WORDS

    )
 
    is_requested = any(rw in cleaned for rw in REQUEST_KEYWORDS)
 
    for bkw in BLOCKER_KEYWORDS:

        if bkw in cleaned:

            blocker_text = cleaned[max(0, cleaned.find(bkw)-20):cleaned.find(bkw)+80].strip()

            return {

                'current_activity': top_act,

                'activity_status': 'Completed with Errors',

                'summary': f"[{email['date'].strftime('%d-%b-%Y')}] Keyword fallback: {blocker_text}",

                'blockers': blocker_text,

                'last_email_date': email['date'],

                'last_email_sender': email.get('sender', ''),

                'email_count': 1,

            }
 
    if activity_completed:

        act_status = 'Completed'

    elif is_requested and top_act not in ('Unknown', 'No Emails'):

        act_status = f'{top_act} Requested'

    elif top_act not in ('Unknown', 'No Emails'):

        act_status = 'In Progress'

    else:

        act_status = 'Unknown'
 
    body_snip = re.sub(r'\s+', ' ', _strip_reply(email.get('body','')[:250])).strip()

    return {

        'current_activity': top_act,

        'activity_status': act_status,

        'summary': f"[{email['date'].strftime('%d-%b-%Y')}] {email.get('sender','')}: {body_snip[:200]}",

        'blockers': None,

        'last_email_date': email['date'],

        'last_email_sender': email.get('sender', ''),

        'email_count': 1,

    }
 
def detect_current_activity(emails):

    if not emails:

        return {'current_activity': 'No Emails', 'activity_status': 'Unknown',

                'summary': 'No emails found for this object.',

                'blockers': None, 'last_email_date': None,

                'last_email_sender': '', 'email_count': 0}

    if LLM_ENABLED:

        try:

            return _detect_via_llm(emails)

        except Exception as e:

            print(f'    LLM detection failed ({e}) — falling back to keyword scoring.')

    return _detect_via_keywords(emails[:1])
 
 
# ═══════════════════════════════════════════════════════════════════════════════

#  SECTION 6 — STATUS COMPARATOR

# ═══════════════════════════════════════════════════════════════════════════════
 
_COMPARE_ORDER = ACTIVITY_ORDER + ['Complete']

RAG_GREEN = 'On Track'

RAG_AMBER = 'At Risk'

RAG_RED   = 'Overdue'

RAG_BLUE  = 'Complete'

RAG_GREY  = 'No Data'
 
 
def _act_idx(act):

    try:    return _COMPARE_ORDER.index(act)

    except: return -1
 
 
def compare_status(obj_data, detected_activity, today=None):

    if today is None: today = date.today()

    expected = get_expected_activity(obj_data, today)
 
    if detected_activity in ('No Emails', 'Unknown', None):

        return {'expected_activity': expected, 'detected_activity': detected_activity or 'Unknown',

                'steps_behind': None, 'days_overdue': None, 'rag_status': RAG_GREY,

                'detail': 'No email data available — status cannot be assessed.'}
 
    if detected_activity == 'Complete' and expected == 'Complete':

        return {'expected_activity': 'Complete', 'detected_activity': 'Complete',

                'steps_behind': 0, 'days_overdue': 0, 'rag_status': RAG_BLUE,

                'detail': 'All migration activities complete.'}
 
    exp_idx, det_idx = _act_idx(expected), _act_idx(detected_activity)

    if exp_idx < 0 or det_idx < 0:

        return {'expected_activity': expected, 'detected_activity': detected_activity,

                'steps_behind': None, 'days_overdue': None, 'rag_status': RAG_GREY,

                'detail': f'Cannot compare: expected={expected}, detected={detected_activity}.'}
 
    steps_behind   = exp_idx - det_idx

    planned_finish = get_planned_finish(obj_data, detected_activity)

    days_overdue   = (today - planned_finish).days if planned_finish else None
 
    if detected_activity == 'Complete':

        rag    = RAG_BLUE;  detail = 'All activities reported complete in emails.'

    elif steps_behind < 0:

        rag    = RAG_GREEN; detail = (f'Ahead of schedule by {abs(steps_behind)} step(s). '

                                      f'At {detected_activity}; POAP expects {expected}.')

    elif steps_behind == 0:

        if days_overdue and days_overdue > 7:

            rag  = RAG_AMBER; detail = (f'On correct activity ({detected_activity}) but '

                                        f'{days_overdue} days past planned finish.')

        else:

            rag  = RAG_GREEN; detail = f'On track. At {detected_activity} (POAP expects {expected}).'

    elif steps_behind == 1:

        if days_overdue and days_overdue > 14:

            rag  = RAG_RED;   detail = (f'1 step behind. At {detected_activity}, expected {expected}. '

                                        f'{detected_activity} is {days_overdue} days overdue.')

        else:

            rag  = RAG_AMBER; detail = f'Slightly behind. At {detected_activity}, POAP expects {expected}.'

    else:

        rag    = RAG_RED

        detail = f'{steps_behind} step(s) behind POAP. At {detected_activity}, expected {expected}.'

        if days_overdue and days_overdue > 0:

            detail += f' {detected_activity} is {days_overdue} days overdue.'
 
    return {'expected_activity': expected, 'detected_activity': detected_activity,

            'steps_behind': steps_behind, 'days_overdue': days_overdue,

            'rag_status': rag, 'detail': detail}
 
 
def get_migration_status(det):

    if not det or det.get('email_count', 0) == 0:

        return 'Not Started'

    act_status = det.get('activity_status', '')

    if act_status == 'Completed':

        return 'Completed'

    if act_status == 'Completed with Errors':

        return 'Completed with Errors'

    if act_status == 'Blocked':

        return 'Completed with Errors'

    if 'Requested' in act_status and not act_status.startswith('Unknown'):

        return 'Requested'

    if act_status == 'Unknown':

        return 'No Data'

    last = det.get('last_email_date')

    if last:

        hours_ago = (datetime.now() - last).total_seconds() / 3600

        if hours_ago <= DUE_HOURS:

            return 'In Progress'

    return 'Due'
 
 
# ═══════════════════════════════════════════════════════════════════════════════

#  SECTION 7 — REPORT WRITER

# ═══════════════════════════════════════════════════════════════════════════════

# (unchanged from your version except Blocked removed from tile colour dicts)
 
_RAG_BG = {'On Track':'C6EFCE','At Risk':'FFEB9C','Overdue':'FFC7CE',

            'Complete':'BDD7EE','No Data':'EDEDED'}

_RAG_FG = {'On Track':'276221','At Risk':'9C6500','Overdue':'9C0006',

            'Complete':'1F4E79','No Data':'595959'}

_THIN   = Border(left=Side(style='thin'), right=Side(style='thin'),

                 top=Side(style='thin'),  bottom=Side(style='thin'))

_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

_WRAP   = Alignment(horizontal='left',   vertical='top',    wrap_text=True)
 
COLUMNS = [

    ('Object ID',10),('Object Name',30),('Stream',20),('Responsible',22),

    ('Start Week',11),('POAP % Done',12),('Detected Activity',20),('Activity Status',16),

    ('Expected Activity',20),('Steps Behind',13),('Days Overdue',13),

    ('Status',14),('Last Email',14),('Emails Read',12),

    ('POAP Load Step',20),('Blockers',40),('Latest Email Summary',65),

]
 
def _xcell(ws, row, col, value, fill=None, font=None, align=None):

    c = ws.cell(row=row, column=col, value=value)

    if fill:  c.fill  = fill

    if font:  c.font  = font

    if align: c.alignment = align

    c.border = _THIN

    return c
 
def write_excel(results):

    os.makedirs(os.path.dirname(OUTPUT_EXCEL) or '.', exist_ok=True)

    wb  = openpyxl.Workbook(); ws = wb.active; ws.title = 'POAP Status Dashboard'

    lc  = get_column_letter(len(COLUMNS))

    now = datetime.now().strftime('%d-%b-%Y %H:%M')
 
    ws.merge_cells(f'A1:{lc}1')

    tc = ws.cell(1, 1, f'R4.0 Mock1 EU  –  POAP Migration Status Monitor  |  {now}')

    tc.fill = PatternFill('solid', fgColor='1F4E79')

    tc.font = Font(bold=True, color='FFFFFF', size=13)

    tc.alignment = _CENTER

    ws.row_dimensions[1].height = 26
 
    for ci, (hdr, width) in enumerate(COLUMNS, 1):

        _xcell(ws, 2, ci, hdr,

               fill=PatternFill('solid', fgColor='2E75B6'),

               font=Font(bold=True, color='FFFFFF', size=11), align=_CENTER)

        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.row_dimensions[2].height = 30
 
    rag_counts = {k: 0 for k in _RAG_BG}

    for ri, r in enumerate(results, 3):

        rag  = r.get('rag_status', 'No Data')

        fill = PatternFill('solid', fgColor=_RAG_BG.get(rag,'FFFFFF'))

        font = Font(color=_RAG_FG.get(rag,'000000'))

        bfnt = Font(color=_RAG_FG.get(rag,'000000'), bold=True)

        rag_counts[rag] = rag_counts.get(rag, 0) + 1
 
        le = r.get('last_email_date')

        le = le.strftime('%d-%b-%Y') if hasattr(le,'strftime') else (le or '—')

        s, o = r.get('steps_behind'), r.get('days_overdue')
 
        vals = [r.get('object_id',''), r.get('name',''), r.get('stream',''),

                r.get('last_email_sender',''), r.get('start_week',''),

                f"{int(r.get('pct_complete',0)*100)}%",

                r.get('detected_activity',''), r.get('migration_status','Not Started'),

                r.get('expected_activity',''),

                s if s is not None else '—', o if o is not None else '—', rag,

                le, r.get('email_count',0), r.get('load_cycle_step',''),

                r.get('blockers','') or '', r.get('summary','')]
 
        for ci, val in enumerate(vals, 1):

            _xcell(ws, ri, ci, val,

                   fill=PatternFill('solid', fgColor=_RAG_BG.get(rag,'FFFFFF')),

                   font=bfnt if ci in (1,10,11) else font,

                   align=_WRAP if ci >= 15 else _CENTER)

        ws.row_dimensions[ri].height = 42
 
    ws.freeze_panes = 'A3'

    ws.auto_filter.ref = f'A2:{lc}{2+len(results)}'
 
    ws2 = wb.create_sheet('RAG Summary')

    ws2.cell(1,1,'Status').font = Font(bold=True)

    ws2.cell(1,2,'Count').font  = Font(bold=True)

    for i,(rag,cnt) in enumerate(rag_counts.items(),2):

        ws2.cell(i,1,rag).fill = PatternFill('solid',fgColor=_RAG_BG.get(rag,'FFFFFF'))

        ws2.cell(i,1).font     = Font(color=_RAG_FG.get(rag,'000000'),bold=True)

        ws2.cell(i,2,cnt)
 
    wb.save(OUTPUT_EXCEL)

    print(f'  Excel  -> {OUTPUT_EXCEL}')

    return rag_counts
 
 
def write_html(results, rag_counts):

    now_str = datetime.now().strftime('%d-%b-%Y %H:%M:%S')

    total   = sum(rag_counts.values())

    HTML_BG = {'On Track':'#C6EFCE','At Risk':'#FFEB9C','Overdue':'#FFC7CE',

               'Complete':'#BDD7EE','No Data':'#EDEDED'}

    HTML_FG = {'On Track':'#276221','At Risk':'#9C6500','Overdue':'#9C0006',

               'Complete':'#1F4E79','No Data':'#595959'}
 
    streams    = sorted({r.get('stream','') for r in results if r.get('stream')})

    activities = sorted({r.get('current_activity','') for r in results

                         if r.get('current_activity') and r.get('current_activity') != 'No Emails'})

    stream_opts   = ''.join(f'<option value="{s}">{s}</option>' for s in streams)

    activity_opts = ''.join(f'<option value="{a}">{a}</option>' for a in activities)
 
    MIG_BG = {'Not Started':'#EDEDED','Requested':'#EDE9FE','In Progress':'#BDD7EE',

              'Due':'#FFEB9C','Completed':'#C6EFCE',

              'Completed with Errors':'#FFD7C4','No Data':'#F0F0F0'}

    MIG_FG = {'Not Started':'#595959','Requested':'#6B21A8','In Progress':'#1F4E79',

              'Due':'#9C6500','Completed':'#276221',

              'Completed with Errors':'#7B3E00','No Data':'#888888'}
 
    mig_counts = {'Not Started':0,'Requested':0,'In Progress':0,'Due':0,'Completed':0,

                  'Completed with Errors':0,'No Data':0}

    for r in results:

        ms = r.get('migration_status','Not Started')

        if ms not in mig_counts:

            mig_counts['No Data'] += 1

        else:

            mig_counts[ms] += 1
 
    tile1_html = (

        f'<div class="card" data-mig="" onclick="migClick(this)" '

        f'style="background:#1F4E79;color:#fff">'

        f'<div class="num">{total}</div><div class="lbl">Total</div></div>'

    )

    _base_tiles = {'Not Started','Requested','In Progress','Due','Completed'}

    for ms, cnt in mig_counts.items():

        if ms not in _base_tiles and cnt == 0:

            continue

        bg = MIG_BG.get(ms, '#EDEDED')

        fg = MIG_FG.get(ms, '#595959')

        tile1_html += (

            f'<div class="card" data-mig="{ms}" onclick="migClick(this)" '

            f'style="background:{bg};color:{fg}">'

            f'<div class="num">{cnt}</div><div class="lbl">{ms}</div></div>'

        )
 
    tile2_html = (

        f'<div class="card" data-rag="" onclick="ragClick(this)" '

        f'style="background:#1F4E79;color:#fff">'

        f'<div class="num">{total}</div><div class="lbl">Total</div></div>'

    )

    for rag, cnt in rag_counts.items():

        tile2_html += (

            f'<div class="card" data-rag="{rag}" onclick="ragClick(this)" '

            f'style="background:{HTML_BG[rag]};color:{HTML_FG[rag]}">'

            f'<div class="num">{cnt}</div><div class="lbl">{rag}</div></div>'

        )
 
    rows_html = ''

    for idx, r in enumerate(results):

        rag    = r.get('rag_status', 'No Data')

        bg, fg = HTML_BG.get(rag, '#fff'), HTML_FG.get(rag, '#000')

        le     = r.get('last_email_date', '')

        if hasattr(le, 'strftime'): le = le.strftime('%d-%b-%Y')

        else: le = le or '—'

        pct     = int(r.get('pct_complete', 0) * 100)

        pct_bar = (f'<div class="pbar-wrap">'

                   f'<div class="pbar"><div class="pbar-in" style="width:{pct}%"></div></div>'

                   f'<span class="pct-lbl">{pct}%</span></div>')

        s, o  = r.get('steps_behind'), r.get('days_overdue')

        sc    = f'<b class="warn">{s}</b>' if (s or 0) > 1 else (str(s) if s is not None else '—')

        oc    = f'<b class="warn">{o}</b>' if (o or 0) > 14 else (str(o) if o is not None else '—')

        blk   = r.get('blockers') or ''

        blk_title = blk.replace('"', '&quot;')

        blk_warn = (f'<span style="color:#dc2626;font-size:15px" title="{blk_title}">&#9888;</span>'

                    if blk else '<span class="na">—</span>')

        cur_act = r.get('current_activity', '') or ''

        stream  = r.get('stream', '') or ''

        summary = (r.get('summary') or '').replace('<', '&lt;').replace('>', '&gt;')

        has_em  = '1' if r.get('last_email_date') else '0'

        has_blk = '1' if blk else '0'

        badge   = (f'<span class="badge" style="background:{HTML_BG[rag]};color:{HTML_FG[rag]}'

                   f';border:1px solid {HTML_FG[rag]}40">{rag}</span>')

        det_blk = (f'<div class="det-blk"><b>Blockers:</b> {blk}</div>' if blk else '')

        det_sum = (f'<div class="det-sum"><b>Latest Summary:</b><br>{summary}</div>'

                   if summary else '<div class="det-sum"><i style="color:#999">No summary yet.</i></div>')
 
        mig      = r.get('migration_status', 'Not Started')

        row_bg   = '#eef4ff' if has_em == '1' else '#ececec'

        row_cls  = 'dr has-em' if has_em == '1' else 'dr no-em'

        id_style = (f'border-left:3px solid {HTML_FG[rag]};font-weight:700;white-space:nowrap'

                    if has_em == '1' else 'border-left:3px solid #e2e8f0;font-weight:700;white-space:nowrap;color:#94a3b8')

        rows_html += (

            f'<tr class="{row_cls}" data-rag="{rag}" data-mig="{mig}" data-stream="{stream}" data-act="{cur_act}" '

            f'data-blk="{has_blk}" data-em="{has_em}" data-idx="{idx}" onclick="toggleDet({idx})" '

            f'style="background:{row_bg}">'

            f'<td style="{id_style}">'

            f'<span class="exp-icon" id="ei{idx}">&#9654;</span>{r.get("object_id","")}</td>'

            f'<td>{r.get("name","")}</td>'

            f'<td>{stream}</td>'

            f'<td class="sm">{r.get("last_email_sender","") or "<span class=na>—</span>"}</td>'

            f'<td class="ctr">{r.get("start_week","")}</td>'

            f'<td>{pct_bar}</td>'

            f'<td class="act">{cur_act}</td>'

            f'<td class="ctr"><span class="badge" style="background:{MIG_BG.get(mig,"#EDEDED")};'

            f'color:{MIG_FG.get(mig,"#595959")};border:1px solid {MIG_FG.get(mig,"#595959")}40">'

            f'{mig}</span></td>'

            f'<td class="exp-act">{r.get("expected_activity","")}</td>'

            f'<td class="ctr">{sc}</td>'

            f'<td class="ctr">{oc}</td>'

            f'<td class="ctr">{badge}</td>'

            f'<td class="ctr sm">{le}</td>'

            f'<td class="ctr" style="padding:8px 4px">{blk_warn}</td>'

            f'</tr>'

            f'<tr class="det-row" id="det{idx}" style="display:none">'

            f'<td colspan="14"><div class="det-box">'

            f'<div class="det-title">'

            f'<span style="color:{fg};font-weight:800;font-size:14px">{r.get("object_id","")} — {r.get("name","")}</span>'

            f'<span class="det-badge" style="background:{HTML_BG[rag]};color:{fg}">{rag}</span>'

            f'<button class="det-close" onclick="event.stopPropagation();toggleDet({idx})">&#10005; Close</button>'

            f'</div>'

            f'<div class="det-grid">'

            f'<div class="det-item"><span class="dk">Object ID</span><span class="dv">{r.get("object_id","")}</span></div>'

            f'<div class="det-item"><span class="dk">Object Name</span><span class="dv">{r.get("name","")}</span></div>'

            f'<div class="det-item"><span class="dk">Stream</span><span class="dv">{stream}</span></div>'

            f'<div class="det-item"><span class="dk">DDME</span><span class="dv">{r.get("owner_eu","")}</span></div>'

            f'<div class="det-item"><span class="dk">POAP Status</span><span class="dv" style="color:{fg};font-weight:700">{rag}</span></div>'

            f'<div class="det-item"><span class="dk">Migration Status</span><span class="dv" style="color:{MIG_FG.get(mig,"#222")};font-weight:700">{mig}</span></div>'

            f'<div class="det-item"><span class="dk">Detected Activity</span><span class="dv">{cur_act}</span></div>'

            f'<div class="det-item"><span class="dk">Expected Activity</span><span class="dv">{r.get("expected_activity","")}</span></div>'

            f'<div class="det-item"><span class="dk">Steps Behind</span><span class="dv">{r.get("steps_behind","—")}</span></div>'

            f'<div class="det-item"><span class="dk">Days Overdue</span><span class="dv">{r.get("days_overdue","—")}</span></div>'

            f'<div class="det-item"><span class="dk">Last Email</span><span class="dv">{le}</span></div>'

            f'<div class="det-item"><span class="dk">POAP Progress</span><span class="dv">{pct}%</span></div>'

            f'<div class="det-item"><span class="dk">Start Week</span><span class="dv">{r.get("start_week","—")}</span></div>'

            f'</div>'

            f'{det_blk}{det_sum}'

            f'</div></td></tr>'

        )

 

    _CSS = """\

:root{--primary:#1e40af;--primary-dark:#1e3a8a;--primary-light:#3b82f6;

      --surface:#fff;--surface-2:#f8fafc;--border:#e2e8f0;

      --text:#1e293b;--text-muted:#64748b;

      --shadow-sm:0 1px 3px rgba(0,0,0,.07),0 1px 2px rgba(0,0,0,.05);

      --shadow-md:0 4px 6px -1px rgba(0,0,0,.08),0 2px 4px -1px rgba(0,0,0,.05);

      --shadow-lg:0 10px 20px -3px rgba(0,0,0,.12),0 4px 8px -2px rgba(0,0,0,.06);

      --radius:14px;--radius-sm:8px;}

*{box-sizing:border-box;margin:0;padding:0}

body{font-family:'Segoe UI Variable','Segoe UI',system-ui,-apple-system,sans-serif;

      background:#f1f5f9;color:var(--text);font-size:13px;line-height:1.5}

.hdr{background:linear-gradient(135deg,#0f2952 0%,#1e40af 55%,#2563eb 100%);

      color:#fff;padding:18px 28px;display:flex;justify-content:space-between;

      align-items:center;position:relative;box-shadow:var(--shadow-md)}

.hdr::after{content:'';position:absolute;bottom:0;left:0;right:0;height:3px;

             background:linear-gradient(90deg,#60a5fa,#818cf8,#a78bfa,#60a5fa)}

.hdr h1{font-size:19px;font-weight:800;letter-spacing:-.3px}

.hdr .sub{font-size:11.5px;opacity:.75;margin-top:4px;font-weight:400}

.hdr .ts{font-size:11px;opacity:.7;text-align:right;line-height:1.8}

.hdr .ts b{display:block;font-size:13px;font-weight:600;opacity:.9}

.tiles{display:flex;gap:16px;padding:20px 28px 14px;flex-wrap:wrap}

.tile{background:var(--surface);border-radius:var(--radius);padding:16px 20px;

       flex:1;min-width:300px;box-shadow:var(--shadow-sm);border:1px solid var(--border)}

.tile-hdr{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.8px;

           color:var(--text-muted);margin-bottom:14px;display:flex;align-items:center;

           gap:7px;padding-bottom:10px;border-bottom:1px solid var(--border)}

.cards{display:flex;gap:10px;flex-wrap:wrap}

.card{border-radius:10px;padding:14px 18px;min-width:96px;text-align:center;

       border:1px solid rgba(0,0,0,.09);cursor:pointer;user-select:none;

       box-shadow:var(--shadow-sm);position:relative;overflow:hidden;

       transition:transform .2s cubic-bezier(.4,0,.2,1),box-shadow .2s cubic-bezier(.4,0,.2,1)}

.card::after{content:'';position:absolute;top:0;left:0;right:0;height:3px;

              background:currentColor;opacity:.35}

.card:hover{transform:translateY(-3px);box-shadow:var(--shadow-md)}

.card.active{transform:translateY(-3px);box-shadow:var(--shadow-lg);

              outline:2.5px solid currentColor;outline-offset:2px}

.num{font-size:30px;font-weight:800;line-height:1;letter-spacing:-1px}

.lbl{font-size:10.5px;font-weight:600;margin-top:5px;opacity:.85;letter-spacing:.2px}

.fbar{padding:10px 28px;display:flex;gap:8px;flex-wrap:wrap;align-items:center;

       background:var(--surface);border-bottom:1px solid var(--border);

       border-top:1px solid var(--border);position:sticky;top:0;z-index:20;

       box-shadow:var(--shadow-sm)}

.fbar input,.fbar select{padding:6px 12px;border:1.5px solid var(--border);

  border-radius:20px;font-size:12px;outline:none;background:var(--surface-2);

  color:var(--text);transition:border-color .15s,box-shadow .15s;font-family:inherit}

.fbar input{width:220px}

.fbar input:focus,.fbar select:focus{border-color:var(--primary-light);

  box-shadow:0 0 0 3px rgba(59,130,246,.12)}

.fbar label{font-size:12px;display:flex;align-items:center;gap:4px;

             cursor:pointer;white-space:nowrap;color:var(--text-muted)}

.fbar label input[type=checkbox]{accent-color:var(--primary);margin:0;padding:0;

  width:13px;height:13px;cursor:pointer;vertical-align:middle}

.fbar .sep{color:var(--border);font-size:20px}

.clr-btn{padding:6px 14px;background:var(--surface-2);border:1.5px solid var(--border);

          border-radius:20px;cursor:pointer;font-size:12px;color:var(--text-muted);

          transition:all .15s;font-family:inherit}

.clr-btn:hover{background:#fee2e2;border-color:#fca5a5;color:#dc2626}

.ref-btn{padding:6px 14px;background:var(--primary);color:#fff;border:none;

          border-radius:20px;cursor:pointer;font-size:12px;font-weight:600;

          transition:background .15s;font-family:inherit}

.ref-btn:hover{background:var(--primary-dark)}

.count-lbl{font-size:12px;color:var(--text-muted);margin-left:auto;white-space:nowrap;font-weight:500}

.wrap{padding:16px 28px 32px;overflow-x:auto}

.tbl-wrap{border-radius:var(--radius);overflow:hidden;box-shadow:var(--shadow-md);

           border:1px solid var(--border);margin-top:4px}

table{width:100%;border-collapse:collapse;background:var(--surface)}

thead{position:sticky;top:0;z-index:10}

thead th{background:#0f2952;color:#cbd5e1;padding:10px 10px;text-align:left;

          font-weight:600;cursor:pointer;user-select:none;white-space:nowrap;

          font-size:11.5px;letter-spacing:.3px;border-right:1px solid rgba(255,255,255,.07)}

thead th:last-child{border-right:none}

thead th:hover{background:#1e3a8a;color:#fff}

.sa{font-size:9px;opacity:.55;margin-left:3px}

tbody td{padding:8px 10px;border-bottom:1px solid #f1f5f9;vertical-align:middle}

tbody tr:last-child td{border-bottom:none}

.dr{cursor:pointer;transition:filter .1s}

.dr:hover td{filter:brightness(.95)}

.no-em td{color:#94a3b8!important}

.no-em .badge{opacity:.6}

.no-em .act{color:#94a3b8!important}

.no-em .pbar-in{background:#cbd5e1}

.det-row{display:none}

.det-box{background:#f8fafc;padding:20px 24px;border-top:3px solid var(--primary-light)}

.det-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(175px,1fr));

           gap:12px;margin-bottom:16px}

.det-item{background:var(--surface);border-radius:var(--radius-sm);padding:11px 14px;

           border:1px solid var(--border);box-shadow:var(--shadow-sm)}

.dk{display:block;font-size:10px;color:var(--text-muted);font-weight:700;

     text-transform:uppercase;letter-spacing:.6px;margin-bottom:3px}

.dv{display:block;font-size:13px;color:var(--text);font-weight:600}

.det-title{display:flex;align-items:center;gap:10px;margin-bottom:16px;

            padding-bottom:12px;border-bottom:1px solid var(--border)}

.det-badge{padding:4px 12px;border-radius:20px;font-size:11px;font-weight:700}

.det-close{margin-left:auto;padding:5px 14px;background:var(--surface);

            border:1.5px solid var(--border);border-radius:20px;cursor:pointer;

            font-size:11px;color:var(--text-muted);transition:all .15s;font-family:inherit}

.det-close:hover{background:#fee2e2;border-color:#fca5a5;color:#dc2626}

.det-blk{background:#fff1f2;border:1px solid #fecdd3;border-radius:var(--radius-sm);

          padding:10px 14px;font-size:12px;color:#be123c;margin-bottom:12px;

          display:flex;gap:8px;align-items:flex-start}

.det-blk::before{content:'\26A0';font-size:14px;flex-shrink:0}

.det-sum{background:var(--surface);border:1px solid var(--border);

          border-radius:var(--radius-sm);border-left:3px solid var(--primary-light);

          padding:12px 16px;font-size:12px;line-height:1.8;color:var(--text)}

.badge{display:inline-flex;align-items:center;gap:5px;padding:3px 10px;

        border-radius:20px;font-size:11px;font-weight:700}

.badge::before{content:'';width:6px;height:6px;border-radius:50%;

                background:currentColor;flex-shrink:0}

.warn{color:#dc2626;font-weight:700}.na{color:#94a3b8}

.blk-txt{color:#dc2626;font-size:11px}

.exp-icon{font-size:9px;margin-right:5px;display:inline-block;

           transition:transform .2s;color:#94a3b8}

.sm{font-size:12px}.ctr{text-align:center}

.act{font-weight:700;color:var(--primary)}.exp-act{color:var(--text-muted);font-size:12px}

.pbar-wrap{display:flex;align-items:center;gap:6px}

.pbar{background:#e2e8f0;border-radius:4px;width:68px;height:5px;flex-shrink:0}

.pbar-in{background:linear-gradient(90deg,var(--primary-light),#818cf8);

          height:5px;border-radius:4px}

.pct-lbl{font-size:11px;color:var(--text-muted);font-weight:600}"""
 
    _JS_TMPL = """\

var sd={}, activeRag='', activeMig='';

var TOTAL=__TOTAL__;

var allDr = Array.from(document.querySelectorAll('.dr'));
 
function updateCount(){

  var v=allDr.filter(function(r){return r.style.display!=='none';}).length;

  document.getElementById('cnt').textContent='Showing '+v+' of '+TOTAL+' objects';

}
 
function applyFilters(){

  var s  =document.getElementById('srch').value.toLowerCase();

  var r  =document.getElementById('rf').value;

  var mf =document.getElementById('mf').value;

  var st =document.getElementById('sf').value;

  var act=document.getElementById('af').value;

  var blk=document.getElementById('blkf').checked;

  var em =document.getElementById('emf').checked;

  allDr.forEach(function(tr){

    var show=(!s  ||tr.textContent.toLowerCase().includes(s))
&&(!r  ||tr.dataset.rag===r)
&&(!mf ||tr.dataset.mig===mf)
&&(!st ||tr.dataset.stream===st)
&&(!act||tr.dataset.act===act)
&&(!blk||tr.dataset.blk==='1')
&&(!em ||tr.dataset.em==='1');

    tr.style.display=show?'':'none';

    var det=document.getElementById('det'+tr.dataset.idx);

    if(det&&!show) det.style.display='none';

  });

  updateCount();

}
 
function ragClick(el){

  var rag=el.getAttribute('data-rag');

  var same=activeRag===rag;

  document.querySelectorAll('#rag-cards .card').forEach(function(c){c.classList.remove('active');});

  if(same){activeRag='';document.getElementById('rf').value='';}

  else{activeRag=rag;document.getElementById('rf').value=rag;el.classList.add('active');}

  applyFilters();

}
 
function migClick(el){

  var mig=el.getAttribute('data-mig');

  var same=activeMig===mig;

  document.querySelectorAll('#mig-cards .card').forEach(function(c){c.classList.remove('active');});

  if(same){activeMig='';document.getElementById('mf').value='';}

  else{activeMig=mig;document.getElementById('mf').value=mig;el.classList.add('active');}

  applyFilters();

}
 
function clearFilters(){

  document.getElementById('srch').value='';

  document.getElementById('rf').value='';

  document.getElementById('mf').value='';

  document.getElementById('sf').value='';

  document.getElementById('af').value='';

  document.getElementById('blkf').checked=false;

  document.getElementById('emf').checked=false;

  activeRag=''; activeMig='';

  document.querySelectorAll('.card').forEach(function(c){c.classList.remove('active');});

  applyFilters();

}
 
function toggleDet(idx){

  var det=document.getElementById('det'+idx);

  var icon=document.getElementById('ei'+idx);

  var open=det.style.display==='table-row';

  det.style.display=open?'none':'table-row';

  icon.style.transform=open?'':'rotate(90deg)';

  icon.style.color=open?'#888':'#2E75B6';

}
 
function srt(c){

  var tb=document.getElementById('tb');

  var ch=Array.from(tb.children);

  var pairs=[];

  for(var i=0;i<ch.length;i+=2) pairs.push([ch[i],ch[i+1]]);

  sd[c]=!sd[c];

  pairs.sort(function(a,b){

    var av=(a[0].cells[c]||{}).textContent.trim();

    var bv=(b[0].cells[c]||{}).textContent.trim();

    var an=parseFloat(av),bn=parseFloat(bv);

    if(!isNaN(an)&&!isNaN(bn)) return sd[c]?an-bn:bn-an;

    return sd[c]?av.localeCompare(bv):bv.localeCompare(av);

  });

  for(var i=0;i<=12;i++){var e=document.getElementById('sa'+i);if(e)e.textContent='';}

  var sa=document.getElementById('sa'+c);if(sa)sa.textContent=sd[c]?' ▲':' ▼';

  pairs.forEach(function(p){tb.appendChild(p[0]);tb.appendChild(p[1]);});

  applyFilters();

}
 
updateCount();"""
 
    js_block = _JS_TMPL.replace('__TOTAL__', str(total))

    excel_base = os.path.basename(OUTPUT_EXCEL)

    html = (

        '<!DOCTYPE html>\n<html lang="en"><head><meta charset="UTF-8">'

        '<meta http-equiv="refresh" content="30">\n'

        '<title>R4.0 - POAP Monitor</title>\n'

        f'<style>\n{_CSS}\n</style></head><body>\n'

        '<div class="hdr">\n  <div>\n'

        '    <h1>R4.0 &ndash; POAP Migration Status Monitor</h1>\n'

        f'    <div class="sub">SAP R4.0 &nbsp;&middot;&nbsp; {total} objects tracked'

        ' &nbsp;&middot;&nbsp; Auto-refreshes every 30s</div>\n'

        '  </div>\n'

        f'  <div class="ts">Last updated<br>{now_str}<br>{excel_base}</div>\n'

        '</div>\n\n<div class="tiles">\n  <div class="tile">\n'

        f'    <div class="tile-hdr">&#9993; Migration Activity &nbsp;'

        f'<span style="font-weight:400;color:#888">(email-based &middot; Due threshold: {DUE_HOURS}h)</span></div>\n'

        f'    <div class="cards" id="mig-cards">{tile1_html}</div>\n'

        '  </div>\n  <div class="tile">\n'

        '    <div class="tile-hdr">&#128203; POAP Plan Status &nbsp;'

        '<span style="font-weight:400;color:#888">(vs planned dates)</span></div>\n'

        f'    <div class="cards" id="rag-cards">{tile2_html}</div>\n'

        '  </div>\n</div>\n\n'

        '<div class="fbar">\n'

        '  <input id="srch" placeholder="&#128269; Search ID / name / owner..."'

        ' oninput="applyFilters()" style="width:220px">\n'

        '  <select id="rf" onchange="applyFilters()">\n'

        '    <option value="">All POAP Status</option>\n'

        '    <option>On Track</option><option>At Risk</option><option>Overdue</option>\n'

        '    <option>Complete</option><option>No Data</option>\n'

        '  </select>\n'

        '  <select id="mf" onchange="applyFilters()">\n'

        '    <option value="">All Activity Status</option>\n'

        '    <option>Not Started</option><option>Requested</option>\n'

        '    <option>In Progress</option><option>Due</option><option>Completed</option>\n'

        '  </select>\n'

        f'  <select id="sf" onchange="applyFilters()"><option value="">All Streams</option>{stream_opts}</select>\n'

        f'  <select id="af" onchange="applyFilters()"><option value="">All Activities</option>{activity_opts}</select>\n'

        '  <span class="sep">|</span>\n'

        '  <label><input type="checkbox" id="blkf" onchange="applyFilters()"> Has Blockers</label>\n'

        '  <label><input type="checkbox" id="emf"  onchange="applyFilters()"> Has Emails</label>\n'

        '  <span class="sep">|</span>\n'

        '  <button class="clr-btn" onclick="clearFilters()">&#10005; Clear</button>\n'

        '  <button class="ref-btn" onclick="window.location.reload()">&#8635; Refresh</button>\n'

        '  <span class="count-lbl" id="cnt"></span>\n'

        '</div>\n\n'

        '<div class="wrap"><div class="tbl-wrap"><table id="tbl">\n'

        '<thead><tr>\n'

        '  <th onclick="srt(0)">ID<span class="sa" id="sa0"></span></th>\n'

        '  <th onclick="srt(1)">Object Name<span class="sa" id="sa1"></span></th>\n'

        '  <th onclick="srt(2)">Stream<span class="sa" id="sa2"></span></th>\n'

        '  <th onclick="srt(3)">Responsible<span class="sa" id="sa3"></span></th>\n'

        '  <th onclick="srt(4)">Week<span class="sa" id="sa4"></span></th>\n'

        '  <th onclick="srt(5)">POAP %<span class="sa" id="sa5"></span></th>\n'

        '  <th onclick="srt(6)">Detected<span class="sa" id="sa6"></span></th>\n'

        '  <th onclick="srt(7)">Activity Status<span class="sa" id="sa7"></span></th>\n'

        '  <th onclick="srt(8)">Expected<span class="sa" id="sa8"></span></th>\n'

        '  <th onclick="srt(9)">Steps<span class="sa" id="sa9"></span></th>\n'

        '  <th onclick="srt(10)">Days<span class="sa" id="sa10"></span></th>\n'

        '  <th onclick="srt(11)">Status<span class="sa" id="sa11"></span></th>\n'

        '  <th onclick="srt(12)">Last Email<span class="sa" id="sa12"></span></th>\n'

        '  <th style="text-align:center;width:36px;padding:10px 6px" title="Blockers">&#9888;</th>\n'

        '</tr></thead>\n'

        f'<tbody id="tb">{rows_html}</tbody>\n'

        '</table></div></div>\n\n'

        f'<script>\n{js_block}\n</script>\n'

        '</body></html>'

    )

    with open(OUTPUT_HTML, 'w', encoding='utf-8') as f:

        f.write(html)

    print(f'  HTML   -> {OUTPUT_HTML}')
 
 
def write_dashboard_config(results):

    def _serial(v):

        if hasattr(v, 'isoformat'):

            return v.isoformat()

        return v
 
    config = {

        'output_file': OUTPUT_EXCEL,

        'last_updated': datetime.now().isoformat(),

        'results': [{k: _serial(v) for k, v in r.items()} for r in results],

    }

    with open(DASHBOARD_CONFIG, 'w', encoding='utf-8') as f:

        json.dump(config, f, indent=2, default=str)

    print(f'  Config -> {DASHBOARD_CONFIG}')
 
# ═══════════════════════════════════════════════════════════════════════════════

#  SECTION 8 — MAIN ORCHESTRATOR

# ═══════════════════════════════════════════════════════════════════════════════
 
def _check_llm():

    if not LLM_ENABLED:

        print('  LLM : Disabled — using keyword scoring.')

        return False

    if not GAIN_API_URL_ANALYSER or not GAIN_API_KEY_ANALYSER:

        print('  LLM : Dify credentials missing. Falling back to keyword scoring.')

        return False

    try:

        import requests, urllib3

        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

        headers = {"Authorization": f"Bearer {GAIN_API_KEY_ANALYSER}", "Content-Type": "application/json"}

        payload = {"inputs": {"query": "ping"}, "response_mode": "blocking", "user": GAIN_USER_ID}

        resp = requests.post(GAIN_API_URL_ANALYSER, json=payload, headers=headers, timeout=10, verify=False)

        resp.raise_for_status()

        print('  LLM : OK — provider=Dify')

        return True

    except Exception as e:

        print(f'  LLM : UNAVAILABLE ({e}) — keyword fallback will be used.')

        return False
 
 
def run_once():

    sep = '=' * 62

    print(f'\n{sep}')

    print(f'  POAP Monitor  |  {datetime.now().strftime("%d-%b-%Y %H:%M:%S")}')

    print(f'{sep}')
 
    _check_llm()
 
    print('\n[1/4] Parsing POAP plan ...')

    objects = load_objects()

    print(f'      {len(objects)} SAP objects loaded.')
 
    print('\n[2/4] Reading Outlook email chains ...')

    known_ids = set(objects.keys())

    last_scan_dt, cached_chains, cached_detections = cache_load()

    cutoff = cache_get_cutoff(last_scan_dt)
 
    if last_scan_dt is None:

        print(f'      First run - full scan (last {LOOKBACK_DAYS} days).')

    else:

        print(f'      Incremental scan from {cutoff.strftime("%d-%b-%Y %H:%M")} '

              f'({CACHE_OVERLAP_MINS}min overlap). Cache has {len(cached_chains)} objects.')
 
    chains      = cached_chains

    updated_oids = set()

    try:

        new_emails   = get_new_emails(known_ids=known_ids, since_dt=cutoff)

        updated_oids = set(new_emails.keys())

        chains       = cache_merge(cached_chains, new_emails)

        new_count    = sum(len(v) for v in new_emails.values())

        found        = sum(1 for v in chains.values() if v)

        print(f'      {new_count} new email(s) fetched. Chains active for {found} / {len(objects)} objects.')

    except ImportError:

        print('      pywin32 not available - using cached data only.')

    except Exception as e:

        print(f'      ERROR reading Outlook: {e}')

        print('      Continuing with cached email data.')
 
    print('\n[3/4] Analysing and comparing ...')

    today = date.today()
 
    def _needs_detection(oid, emails):

        cached_det   = cached_detections.get(oid)

        chain_latest = emails[0]['date'] if emails else None

        cached_ts    = cached_det.get('last_email_date') if cached_det else None

        cache_stale  = (

            chain_latest and cached_ts and

            isinstance(chain_latest, datetime) and isinstance(cached_ts, datetime) and

            chain_latest > cached_ts

        )

        if cached_det and oid not in updated_oids and not cache_stale:

            return False, cached_det

        return True, None
 
    cached_dets   = {}

    needs_llm     = {}

    for oid, obj_data in objects.items():

        emails = chains.get(oid, [])

        fresh, det = _needs_detection(oid, emails)

        if fresh:

            needs_llm[oid] = emails

        else:

            cached_dets[oid] = det
 
    llm_count = len(needs_llm)

    print(f'      {llm_count} object(s) need LLM detection, {len(cached_dets)} using cache.')
 
    llm_results = {}

    if needs_llm:

        with ThreadPoolExecutor(max_workers=5) as pool:

            future_to_oid = {pool.submit(detect_current_activity, emails): oid

                             for oid, emails in needs_llm.items()}

            for future in as_completed(future_to_oid):

                oid = future_to_oid[future]

                try:

                    llm_results[oid] = future.result()

                except Exception as e:

                    print(f'      ERROR detecting {oid}: {e}')

                    llm_results[oid] = {'current_activity': 'Unknown', 'activity_status': 'Unknown',

                                        'summary': 'Detection failed.', 'blockers': None,

                                        'last_email_date': None, 'last_email_sender': '', 'email_count': 0}
 
    detections = {**cached_dets, **llm_results}

    results = []

    for oid, obj_data in objects.items():

        det = detections.get(oid, {'current_activity': 'No Emails', 'activity_status': 'Unknown',

                                   'summary': '', 'blockers': None, 'last_email_date': None,

                                   'last_email_sender': '', 'email_count': 0})

        if obj_data.get('load_cycle_step') == 'Completed' and det.get('current_activity') not in ('Complete', 'No Emails'):

            det = dict(det)

            det['current_activity'] = 'Complete'

            if not det.get('activity_status', '').startswith('Completed'):

                det['activity_status'] = 'Completed'

        detections[oid] = det

        comp = compare_status(obj_data, det['current_activity'], today)

        results.append({

            'object_id': oid,   'name': obj_data['name'],

            'stream': obj_data['stream'],   'owner_eu': obj_data['owner_eu'],

            'owner_gdme': obj_data['owner_gdme'],   'start_week': obj_data['start_week'],

            'pct_complete': obj_data['pct_complete'],

            'load_cycle_step': obj_data['load_cycle_step'],

            **{k: det[k]  for k in ('current_activity','activity_status','summary',

                                     'blockers','last_email_date','last_email_sender','email_count')},

            **{k: comp[k] for k in ('expected_activity','steps_behind',

                                     'days_overdue','rag_status','detail')},

            'migration_status': get_migration_status(det),

        })
 
    _rag_order = {'Overdue':0,'At Risk':1,'No Data':2,'On Track':3,'Complete':4}

    results.sort(key=lambda r: (_rag_order.get(r['rag_status'],9), r.get('start_week',''), r['object_id']))
 
    cache_save(chains, detections)
 
    print('\n[4/4] Writing dashboard ...')

    os.makedirs(STATE_DIR, exist_ok=True)

    rag_counts = write_excel(results)

    write_html(results, rag_counts)

    write_dashboard_config(results)
 
    summary = '  |  '.join(f'{k}: {v}' for k, v in rag_counts.items() if v > 0)

    print(f'\n  {len(results)} objects  |  {summary}')

    print(f'  Dashboard: {OUTPUT_HTML}')
 
    with open(STATE_FILE, 'w') as f:

        json.dump({'last_run': datetime.now().isoformat(), 'object_count': len(results),

                   'rag_counts': rag_counts, 'email_chains': len(chains)}, f, indent=2)

    return rag_counts
 
 
def main():

    parser = argparse.ArgumentParser(description='POAP Migration Monitor')

    parser.add_argument('--once',     action='store_true', help='Run once and exit')

    parser.add_argument('--interval', type=int, default=POLL_SECONDS,

                        help=f'Polling interval in seconds (default: {POLL_SECONDS})')

    args = parser.parse_args()
 
    if args.once:

        run_once(); return
 
    print(f'POAP Monitor started - polling every {args.interval}s.  Ctrl+C to stop.')

    while True:

        try:

            run_once()

        except KeyboardInterrupt:

            print('\nStopped.'); break

        except Exception as e:

            import traceback

            print(f'\n  ERROR: {e}'); traceback.print_exc()

        print(f'\n  Next run in {args.interval}s ...\n')

        try:

            time.sleep(args.interval)

        except KeyboardInterrupt:

            print('\nStopped.'); break
 
 
if __name__ == '__main__':

    main()
 